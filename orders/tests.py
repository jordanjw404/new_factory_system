# orders/tests.py
from datetime import date
from io import BytesIO

import openpyxl
from django.contrib.auth.models import User
from django.test import TestCase
from django.urls import reverse
from customers.models import Customer
from orders.models import Order


class OrderViewsTests(TestCase):
    def setUp(self):
        self.user = User.objects.create_user(username="testuser", password="password123")
        self.client.login(username="testuser", password="password123")
        self.customer = Customer.objects.create(name="Acme Co.")

    def test_order_list_view(self):
        url = reverse("orders:order_list")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<table")

    def test_order_detail_view(self):
        order = Order.objects.create(
            name="My Order",
            customer=self.customer,
            reference="REF123",
            delivery_date=date.today(),
            is_collection=False,
            order_type=Order.OrderType.KITCHEN,
            status=Order.Status.NO_PAPERWORK,
            priority=Order.Priority.MEDIUM,
            robes=1,
            cabs=2,
            panels=3,
            send_to_production=False,
            owner=self.user,
            created_by=self.user,
        )
        url = reverse("orders:order_detail", args=[order.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "My Order")
        self.assertContains(response, "REF123")

    def test_order_create_view_get(self):
        url = reverse("orders:order_create")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<form")

    def test_order_create_view_post_valid(self):
        url = reverse("orders:order_create")
        data = {
            "name": "New Order",
            "customer": self.customer.pk,
            "reference": "NEWREF",
            "delivery_date": date.today().isoformat(),
            "is_collection": False,
            "order_type": Order.OrderType.KITCHEN,
            "status": Order.Status.NO_PAPERWORK,
            "priority": Order.Priority.MEDIUM,
            "robes": 0,
            "cabs": 1,
            "panels": 0,
            "send_to_production": False,
        }
        response = self.client.post(url, data)
        self.assertEqual(response.status_code, 302)
        self.assertTrue(Order.objects.filter(reference="NEWREF").exists())

    def test_order_detail_list_view(self):
        url = reverse("orders:order_detail_list")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<form")

    def test_export_orders_excel(self):
        # create an order so export has content
        Order.objects.create(
            name="XOrder",
            customer=self.customer,
            reference="XREF",
            delivery_date=date.today(),
            is_collection=False,
            order_type=Order.OrderType.KITCHEN,
            status=Order.Status.NO_PAPERWORK,
            priority=Order.Priority.MEDIUM,
            robes=0,
            cabs=1,
            panels=0,
            send_to_production=False,
            owner=self.user,
            created_by=self.user,
        )
        url = reverse("orders:order_export")
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=1, column=1).value, "ID")
        self.assertEqual(ws.cell(row=1, column=2).value, "Customer")

    def test_order_edit_view_get(self):
        order = Order.objects.create(
            name="EditMe",
            customer=self.customer,
            reference="EDITREF",
            delivery_date=date.today(),
            is_collection=False,
            order_type=Order.OrderType.KITCHEN,
            status=Order.Status.NO_PAPERWORK,
            priority=Order.Priority.MEDIUM,
            robes=1,
            cabs=1,
            panels=0,
            send_to_production=False,
            owner=self.user,
            created_by=self.user,
        )
        url = reverse("orders:order_edit", args=[order.pk])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "<form")

    def test_order_delete_view_post(self):
        order = Order.objects.create(
            name="DelMe",
            customer=self.customer,
            reference="DELREF",
            delivery_date=date.today(),
            is_collection=False,
            order_type=Order.OrderType.KITCHEN,
            status=Order.Status.NO_PAPERWORK,
            priority=Order.Priority.MEDIUM,
            robes=0,
            cabs=1,
            panels=0,
            send_to_production=False,
            owner=self.user,
            created_by=self.user,
        )
        url = reverse("orders:order_delete", args=[order.pk])
        response = self.client.post(url)
        self.assertRedirects(response, reverse("orders:order_list"))
        self.assertFalse(Order.objects.filter(pk=order.pk).exists())
